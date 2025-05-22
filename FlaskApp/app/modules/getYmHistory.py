"""
YouTube視聴履歴とチャンネル情報をExcelファイルに保存するスクリプト

このスクリプトは、YouTubeの視聴履歴やチャンネル情報を含むJSONファイルを読み込み、
指定されたExcelファイルに保存します。YouTube Musicなどの特定のカテゴリの視聴履歴や、
YouTubeチャンネルの情報をExcel形式で整理することが可能です。

ディレクトリ構造:
    - JSONファイルは、'json' ディレクトリに保存されている必要があります。
    - 生成されたExcelファイルは、'excel' ディレクトリに保存されます。

関数:
    - getYouTubeHistory(json_file_name, excel_file_name, header_options=["YouTube Music"]):
        JSONファイルから指定されたカテゴリ（デフォルトではYouTube Music）の視聴履歴を
        フィルタリングし、Excelファイルに保存します。視聴タイトル、チャンネル名、視聴日時、
        URLなどの情報を整形して出力します。

    - getChannelId(json_file_name, excel_file_name, header="YouTube"):
        YouTubeチャンネル情報を含むJSONファイルを読み込み、チャンネル名やURL、
        チャンネルIDをExcelファイルに整理して保存します。

使用方法:
    - このスクリプトは、`__main__` セクション内で実行される例を提供しています。
    - 具体的には、`getYouTubeHistory` 関数で視聴履歴を読み込み、
      'youtube_music_watch_history.xlsx' に保存する手順が示されています。
      ファイル名やディレクトリが存在しない場合はエラーメッセージが表示されます。

例:
    python script.py

注意事項:
    - JSONファイルが適切なフォーマットでない場合、エラーが発生する可能性があります。
    - `getYouTubeHistory` 関数では、日本時間（UTC+9）に変換して視聴日時を保存します。
    - Excelファイルが存在しない場合は新規作成し、既存ファイルが存在する場合はシートを
      追加または更新します。

依存パッケージ:
    - pandas: データフレーム操作用
    - openpyxl: Excel操作用
    - json: JSONファイルの読み込み用
    - os: ファイルとディレクトリの操作用
    - datetime: 日付・時間操作用

"""


import os
import json
import requests
from datetime import datetime, timedelta

import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Alignment
from sklearn.cluster import DBSCAN


# ディレクトリのパスを設定
EXCEL_DIR = 'excel'
JSON_DIR = 'json'
# START_DATE = "2025-3-1"
START_DATE = "2025-4-6"

def getYouTubeHistory(json_file_name, excel_file_name, header_options=["YouTube Music"]):
    # JSONファイルのフルパス
    json_file_path = os.path.join(JSON_DIR, json_file_name)

    # JSONデータの読み込み
    try:
        with open(json_file_path, 'r', encoding='utf-8') as f:
            data = json.load(f)
    except FileNotFoundError:
        print(f"JSONファイル '{json_file_path}' が見つかりません。")
        return
    except json.JSONDecodeError:
        print(f"JSONファイル '{json_file_path}' の読み込み中にエラーが発生しました。")
        return

    # データの整形
    watch_history = []

    for entry in data:
        if entry.get('header') in header_options and 'title' in entry and 'time' in entry:
            title = entry['title'].replace(' を視聴しました', '')
            timestamp = entry['time'].replace('Z', '')

            title_url = entry['titleUrl'].replace('\\u003d', '=')
            subtitles_names = [subtitle['name'] for subtitle in entry.get('subtitles', [])]
            subtitles_names_str = ', '.join(subtitles_names)

            watch_history.append({
                'Title': title,
                'Channel': subtitles_names_str,
                'Date': timestamp,
                'Url': title_url,
                'Rank': '',
                'Note': ''
            })

    # DataFrameの作成
    df = pd.DataFrame(watch_history)
    df['Date'] = pd.to_datetime(df['Date'], errors='coerce') + pd.Timedelta(hours=9)
    df['Date'] = df['Date'].dt.strftime('%Y-%m-%d %H:%M:%S')

    # 現在の日本時間を取得してシート名に使用
    now = pd.Timestamp.now(tz='Asia/Tokyo')
    sheet_name = now.strftime('%Y%m%d%H%M')

    # Excelファイルのフルパス
    excel_file_path = os.path.join(EXCEL_DIR, excel_file_name)

    try:
        if not os.path.exists(excel_file_path):
            wb = Workbook()
            ws = wb.active
            ws.title = sheet_name

            for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
                for c_idx, value in enumerate(row, 1):
                    ws.cell(row=r_idx, column=c_idx, value=value)

            for r_idx in range(2, len(df) + 2):
                url_cell = ws.cell(row=r_idx, column=4)  # 4列目にUrlがある
                url_cell.hyperlink = df.iloc[r_idx-2]['Url']  # 'Url'を参照
                url_cell.value = df.iloc[r_idx-2]['Url']
                url_cell.style = 'Hyperlink'

            for col in ws.columns:
                for cell in col:
                    cell.alignment = Alignment(horizontal='center')

            wb.save(excel_file_path)
            print(f"新しいファイルを作成しました: '{excel_file_path}'")
        else:
            with pd.ExcelWriter(excel_file_path, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
                df.to_excel(writer, sheet_name=sheet_name, index=False)
                for r_idx in range(2, len(df) + 2):
                    url_cell = writer.sheets[sheet_name].cell(row=r_idx, column=4)  # 4列目にUrlがある
                    url_cell.hyperlink = df.iloc[r_idx-2]['Url']  # 'Url'を参照
                    url_cell.value = df.iloc[r_idx-2]['Url']
                    url_cell.style = 'Hyperlink'
            print(f"既存のファイルに追加しました: '{excel_file_path}'")

    except Exception as e:
        print(f"Excelファイルの作成中にエラーが発生しました: {e}")

def getChannelId(json_file_name, excel_file_name, header="YouTube"):
    json_file_path = os.path.join(JSON_DIR, json_file_name)

    try:
        with open(json_file_path, 'r', encoding='utf-8') as f:
            data = json.load(f)
    except FileNotFoundError:
        print(f"JSONファイル '{json_file_path}' が見つかりません。")
        return
    except json.JSONDecodeError:
        print(f"JSONファイル '{json_file_path}' の読み込み中にエラーが発生しました。")
        return

    youtube_data = []

    for entry in data:
        if entry.get('header') == header and 'subtitles' in entry:
            for subtitle in entry['subtitles']:
                name = subtitle.get('name', '')
                url = subtitle.get('url', '')
                channel_id = url.split('/')[-1] if url else ''
                youtube_data.append({
                    'Name': name,
                    'URL': url,
                    'Channel ID': channel_id
                })

    df = pd.DataFrame(youtube_data)
    now = pd.Timestamp.now(tz='Asia/Tokyo')
    sheet_name = now.strftime('%Y%m%d%H%M')

    excel_file_path = os.path.join(EXCEL_DIR, excel_file_name)

    try:
        if not os.path.exists(excel_file_path):
            wb = Workbook()
            ws = wb.active
            ws.title = sheet_name

            for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
                for c_idx, value in enumerate(row, 1):
                    ws.cell(row=r_idx, column=c_idx, value=value)

            for r_idx in range(2, len(df) + 2):
                url_cell = ws.cell(row=r_idx, column=2)
                url_cell.hyperlink = df.iloc[r_idx-2]['URL']
                url_cell.value = df.iloc[r_idx-2]['URL']
                url_cell.style = 'Hyperlink'

            for col in ws.columns:
                for cell in col:
                    cell.alignment = Alignment(horizontal='center')

            wb.save(excel_file_path)
            print(f"新しいファイルを作成しました: '{excel_file_path}'")
        else:
            with pd.ExcelWriter(excel_file_path, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
                df.to_excel(writer, sheet_name=sheet_name, index=False)
                for r_idx in range(2, len(df) + 2):
                    url_cell = writer.sheets[sheet_name].cell(row=r_idx, column=2)
                    url_cell.hyperlink = df.iloc[r_idx-2]['URL']
                    url_cell.value = df.iloc[r_idx-2]['URL']
                    url_cell.style = 'Hyperlink'
            print(f"既存のファイルに追加しました: '{excel_file_path}'")

    except Exception as e:
        print(f"Excelファイルの作成中にエラーが発生しました: {e}")

def cluster_youtube_data(file_path, start_date=START_DATE, end_date=None, eps=300, min_samples=5):
    # Excelファイルを読み込む
    data = pd.read_excel(file_path, sheet_name=None)
    
    # 一番右のシート名を取得
    sheet_names = list(data.keys())
    latest_sheet_name = sheet_names[-1]
    df = data[latest_sheet_name]

    # 時系列データが含まれる列名を指定
    time_column = 'Date'  # 時間の列名を指定
    title_column = 'Title'  # タイトルの列名を指定
    channel_column = 'Channel'  # チャンネルの列名を指定

    # 日付列を datetime 型に変換
    df[time_column] = pd.to_datetime(df[time_column])

    # 日付の範囲でフィルタリング
    if start_date:
        start_date = pd.to_datetime(start_date)
        df = df[df[time_column] >= start_date]
    if end_date:
        end_date = pd.to_datetime(end_date)
        df = df[df[time_column] <= end_date]

    # タイムスタンプを計算
    df['Timestamp'] = df[time_column].astype(np.int64) // 10**9  # 秒単位のタイムスタンプ

    # DBSCANクラスタリングを実行
    dbscan = DBSCAN(eps=eps, min_samples=min_samples)
    df['Cluster'] = dbscan.fit_predict(df[['Timestamp']])

    # クラスタごとのデータを整理
    clusters_info = []
    for cluster in np.unique(df['Cluster']):
        if cluster != -1:  # -1はノイズとする
            cluster_data = df[df['Cluster'] == cluster]
            start_date = cluster_data[time_column].min().strftime('%Y-%m-%d %H:%M:%S')
            end_date = cluster_data[time_column].max().strftime('%Y-%m-%d %H:%M:%S')
            sample_count = cluster_data.shape[0]

            # タイトルとチャンネルの組み合わせを取得（最大5つ）
            samples = cluster_data[[title_column, channel_column]].head(5).to_dict(orient='records')

            cluster_dict = {
                'Cluster': int(cluster),
                'Start Date': start_date,
                'End Date': end_date,
                'Sample Count': sample_count,
                'Samples': samples
            }
            clusters_info.append(cluster_dict)

    return clusters_info

def send_to_gas_ym_history(file_path='excel/youtube_music_watch_history.xlsx'):
    # 使用例
    url = os.getenv('GAS_YM_URL')
    result = cluster_youtube_data(file_path)

    # データをJSON形式に変換
    json_data = json.dumps(result)

    # リクエストヘッダーを設定
    headers = {
        'Content-Type': 'application/json',
    }

    # POSTリクエストを送信
    response = requests.post(url, data=json_data, headers=headers)  # gas_endpoint を url に修正

    # レスポンスを確認
    if response.status_code == 200:
        print('データが正常に送信されました:', response.json())
    else:
        print('エラーが発生しました:', response.status_code, response.text)

if __name__ == '__main__':
    # 使用例
    getYouTubeHistory('watch-history.json', 'youtube_music_watch_history.xlsx')
    send_to_gas_ym_history()
